import openpyxl

class ExcelReader:
    @staticmethod
    def get_data(file_path, row_index):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Создаем словарь для удобного доступа к данным
        data = {
            "origin": sheet.cell(row=row_index, column=1).value,
            "destination": sheet.cell(row=row_index, column=2).value,
            "email": sheet.cell(row=row_index, column=3).value,
            "phone": str(sheet.cell(row=row_index, column=4).value),
            "name": sheet.cell(row=row_index, column=5).value,
            "lastname": sheet.cell(row=row_index, column=6).value,
            "passport": str(sheet.cell(row=row_index, column=7).value),
            "nationality": sheet.cell(row=row_index, column=8).value,
        }
        return data